"""
ExcelExporter – converts parsed PLC records into a formatted Excel workbook.

Sheet structure:
  - "Records"      : main decoded data table (one row per record)
  - "Registers"    : pivoted register view
  - "Packet Stats" : summary statistics
  - "Error Log"    : failed/corrupt packets
  - "Raw Hex"      : hex dump per record

Applies professional formatting:
  - Frozen header row
  - Auto-column widths
  - Color-coded validity
  - LiraNET protocol header block
"""

from pathlib import Path
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

from parser.base_parser import ParsedRecord, ParseStats
from utils.logger import AppLogger

log = AppLogger()

# ---------- Color palette ----------
CLR_HEADER_BG = "1E2D40"       # Dark navy (header)
CLR_HEADER_FG = "FFFFFF"       # White text
CLR_VALID_BG = "E8F5E9"        # Light green (valid row)
CLR_INVALID_BG = "FFEBEE"      # Light red (invalid row)
CLR_ACCENT = "2196F3"          # Blue accent
CLR_SUBHEADER_BG = "263850"    # Slightly lighter navy
CLR_EVEN_ROW = "F7F9FC"        # Alternating row color
CLR_BORDER = "B0BEC5"          # Border color


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _make_border() -> Border:
    side = Side(style="thin", color=CLR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


class ExcelExporter:
    """
    Exports parsed PLC records to a fully-formatted .xlsx workbook.
    """

    def __init__(self, output_path: str | Path, config: dict):
        self.output_path = Path(output_path)
        self.config = config
        self._wb = openpyxl.Workbook()
        self._wb.remove(self._wb.active)  # remove default sheet

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export(
        self,
        records: list[ParsedRecord],
        stats: ParseStats,
        errors: list[dict],
        source_file: str = "",
    ) -> Path:
        """
        Build and save the Excel workbook.

        Args:
            records:     Parsed records from HexParser
            stats:       ParseStats summary
            errors:      List of error dicts from parser
            source_file: Name of the source data file (for metadata)
        Returns:
            Path to saved .xlsx file
        """
        log.info("Exporting %d records to Excel: %s", len(records), self.output_path)

        self._add_records_sheet(records)
        self._add_registers_sheet(records)
        self._add_stats_sheet(stats, source_file)
        self._add_error_log_sheet(errors)
        self._add_hex_dump_sheet(records)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(self.output_path)
        log.info("Excel saved: %s", self.output_path)
        return self.output_path

    # ------------------------------------------------------------------
    # Sheet 1: Records
    # ------------------------------------------------------------------

    def _add_records_sheet(self, records: list[ParsedRecord]):
        ws = self._wb.create_sheet("Records")

        # Protocol header
        self._write_protocol_header(ws, col_span=14)

        # Build DataFrame
        rows = [r.to_flat_dict() for r in records]
        if not rows:
            ws.append(["No records decoded."])
            return
        df = pd.DataFrame(rows)
        df = df.fillna("")

        # Write column headers at row 5
        header_row = 5
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = _make_font(bold=True, color=CLR_HEADER_FG, size=10)
            cell.fill = _make_fill(CLR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = _make_border()

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # Write data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
            is_valid = str(getattr(row, "Valid", "✓")) == "✓"
            row_fill = _make_fill(CLR_VALID_BG if is_valid else CLR_INVALID_BG)
            alt_fill = _make_fill(CLR_EVEN_ROW)
            use_fill = row_fill if not is_valid else (
                alt_fill if (row_idx - header_row) % 2 == 0 else None
            )
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = _make_font(size=9)
                cell.border = _make_border()
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if use_fill:
                    cell.fill = use_fill

        # Auto-size columns
        self._auto_size_columns(ws, min_row=header_row, max_col=len(df.columns))

        # Add Excel table
        if len(records) > 0:
            table_ref = (
                f"A{header_row}:{get_column_letter(len(df.columns))}"
                f"{header_row + len(records)}"
            )
            table = Table(displayName="RecordsTable", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            ws.add_table(table)

    # ------------------------------------------------------------------
    # Sheet 2: Registers (pivot view)
    # ------------------------------------------------------------------

    def _add_registers_sheet(self, records: list[ParsedRecord]):
        ws = self._wb.create_sheet("Registers")
        self._write_protocol_header(ws, col_span=8)

        valid_records = [r for r in records if r.is_valid]
        if not valid_records:
            ws.append(["No valid records to display."])
            return

        # Collect all unique register addresses
        all_regs: set[str] = set()
        for r in valid_records:
            all_regs.update(r.registers.keys())
        reg_list = sorted(all_regs)

        # Build config-based description map
        field_defs = self.config.get("fields", [])
        reg_desc = {f.get("register", ""): f.get("description", "") for f in field_defs}
        reg_name = {f.get("register", ""): f.get("name", "") for f in field_defs}

        header_row = 5
        headers = ["Record #", "Timestamp"] + reg_list
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = _make_font(bold=True, color=CLR_HEADER_FG)
            cell.fill = _make_fill(CLR_HEADER_BG)
            cell.alignment = Alignment(horizontal="center")
            cell.border = _make_border()

        # Description row
        desc_row = header_row + 1
        ws.cell(row=desc_row, column=1, value="Description →")
        ws.cell(row=desc_row, column=1).font = _make_font(bold=True, size=8)
        for col_idx, reg in enumerate(reg_list, start=3):
            cell = ws.cell(row=desc_row, column=col_idx,
                           value=reg_desc.get(reg, reg_name.get(reg, "")))
            cell.font = _make_font(size=8, color="555555")
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = ws.cell(row=desc_row + 1, column=1)

        # Data rows
        for row_idx, record in enumerate(valid_records, start=desc_row + 1):
            ws.cell(row=row_idx, column=1, value=record.record_index)
            ws.cell(row=row_idx, column=2, value=record.timestamp_str)
            for col_idx, reg in enumerate(reg_list, start=3):
                reg_info = record.registers.get(reg, {})
                val = reg_info.get("decimal")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = _make_font(size=9)
                cell.border = _make_border()
                if (row_idx - desc_row) % 2 == 0:
                    cell.fill = _make_fill(CLR_EVEN_ROW)

        self._auto_size_columns(ws, min_row=header_row, max_col=len(headers))

    # ------------------------------------------------------------------
    # Sheet 3: Statistics
    # ------------------------------------------------------------------

    def _add_stats_sheet(self, stats: ParseStats, source_file: str):
        ws = self._wb.create_sheet("Packet Stats")

        def _write_kv(row: int, key: str, value: Any):
            kc = ws.cell(row=row, column=1, value=key)
            kc.font = _make_font(bold=True, size=10)
            kc.fill = _make_fill(CLR_SUBHEADER_BG)
            kc.font = Font(bold=True, color=CLR_HEADER_FG, size=10, name="Calibri")
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = _make_font(size=10)

        rows_data = [
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source File", source_file),
            ("", ""),
            ("Total Packets", stats.total_packets),
            ("Valid Packets", stats.valid_packets),
            ("Invalid Packets", stats.invalid_packets),
            ("Success Rate (%)", f"{stats.success_rate:.1f}%"),
            ("Total Data (bytes)", stats.total_bytes),
            ("Parse Duration (s)", f"{stats.parse_duration_s:.3f}"),
            ("", ""),
            ("Timestamp Start", stats.timestamp_range[0]),
            ("Timestamp End", stats.timestamp_range[1]),
            ("", ""),
        ]

        if stats.error_summary:
            rows_data.append(("--- Error Summary ---", "Count"))
            for err, count in sorted(stats.error_summary.items(), key=lambda x: -x[1]):
                rows_data.append((err, count))

        title_cell = ws.cell(row=1, column=1, value="PLC Decoder — Parse Statistics")
        title_cell.font = Font(bold=True, size=14, color=CLR_HEADER_FG, name="Calibri")
        title_cell.fill = _make_fill(CLR_HEADER_BG)
        ws.merge_cells("A1:B1")

        for r_idx, (k, v) in enumerate(rows_data, start=3):
            _write_kv(r_idx, k, v)

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32

    # ------------------------------------------------------------------
    # Sheet 4: Error Log
    # ------------------------------------------------------------------

    def _add_error_log_sheet(self, errors: list[dict]):
        ws = self._wb.create_sheet("Error Log")
        headers = ["#", "Offset (bytes)", "Error", "Raw Hex (first 64 chars)"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = _make_font(bold=True, color=CLR_HEADER_FG)
            cell.fill = _make_fill(CLR_HEADER_BG)
            cell.border = _make_border()

        ws.freeze_panes = "A2"
        for row_idx, err in enumerate(errors, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=err.get("offset", ""))
            ws.cell(row=row_idx, column=3, value=err.get("error", ""))
            cell = ws.cell(row=row_idx, column=4, value=err.get("raw_hex", "")[:64])
            cell.font = Font(name="Courier New", size=8)
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = _make_border()
                ws.cell(row=row_idx, column=col).fill = _make_fill(CLR_INVALID_BG)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 72

    # ------------------------------------------------------------------
    # Sheet 5: Hex Dump
    # ------------------------------------------------------------------

    def _add_hex_dump_sheet(self, records: list[ParsedRecord]):
        ws = self._wb.create_sheet("Raw Hex")
        ws.cell(row=1, column=1, value="Record #").font = _make_font(bold=True)
        ws.cell(row=1, column=2, value="Timestamp").font = _make_font(bold=True)
        ws.cell(row=1, column=3, value="Raw Hex (word-spaced)").font = _make_font(bold=True)
        for col in range(1, 4):
            ws.cell(row=1, column=col).fill = _make_fill(CLR_HEADER_BG)
            ws.cell(row=1, column=col).font = _make_font(bold=True, color=CLR_HEADER_FG)

        ws.freeze_panes = "A2"
        for row_idx, record in enumerate(records, start=2):
            ws.cell(row=row_idx, column=1, value=record.record_index)
            ws.cell(row=row_idx, column=2, value=record.timestamp_str)
            cell = ws.cell(row=row_idx, column=3, value=record.raw_hex_spaced)
            cell.font = Font(name="Courier New", size=8)
            if row_idx % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=row_idx, column=col).fill = _make_fill(CLR_EVEN_ROW)

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 120

    # ------------------------------------------------------------------
    # Utilities
    # ------------------------------------------------------------------

    def _write_protocol_header(self, ws, col_span: int = 10):
        """Write the LiraNET-style protocol header block (rows 1–3)."""
        cfg_meta = self.config.get("metadata", {})
        protocol = cfg_meta.get("protocol_name", "LiraNET Data Transmission Protocol on EtherNET")
        copyright_str = cfg_meta.get("copyright", "mrpl@2016")
        generated_str = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Row 1: Title + copyright
        title_cell = ws.cell(row=1, column=3, value=protocol)
        title_cell.font = Font(bold=True, size=12, color=CLR_HEADER_FG, name="Calibri")
        title_cell.fill = _make_fill(CLR_HEADER_BG)

        copy_cell = ws.cell(row=1, column=col_span, value=copyright_str)
        copy_cell.font = Font(italic=True, size=10, color=CLR_HEADER_FG)
        copy_cell.fill = _make_fill(CLR_HEADER_BG)
        copy_cell.alignment = Alignment(horizontal="right")

        ws.merge_cells(f"A1:B1")
        ws.cell(row=1, column=1).fill = _make_fill(CLR_HEADER_BG)
        ws.cell(row=1, column=2).fill = _make_fill(CLR_HEADER_BG)

        # Row 2: empty spacer with bg
        for col in range(1, col_span + 1):
            ws.cell(row=2, column=col).fill = _make_fill(CLR_SUBHEADER_BG)

        # Row 3: InputMessage + generated date
        ws.cell(row=3, column=3, value="Decoded PLC Records").font = Font(
            bold=True, size=10, color="2196F3"
        )
        ws.cell(row=3, column=col_span, value=generated_str).font = Font(
            size=9, italic=True, color="777777"
        )
        ws.cell(row=3, column=3).alignment = Alignment(horizontal="left")

        # Row 4: spacer
        ws.row_dimensions[4].height = 6

    @staticmethod
    def _auto_size_columns(ws, min_row: int = 1, max_col: int = 20):
        """Auto-fit column widths based on content, capped at 60 chars."""
        for col_idx in range(1, max_col + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=min_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        cell_len = len(str(cell.value)) if cell.value else 0
                        max_len = max(max_len, cell_len)
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
